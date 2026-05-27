"""Merge data_cache/user_input.json → portfolio.xlsx (Positions + Notes sheet).

Source: 用户在 dashboard 编辑 inline form 后点 "💾 导出 user_input.json" 下载的
JSON, 用户自己放到 data_cache/user_input.json (单一 well-known 路径).

JSON schema:
    {
      "SH600547__buy_price": "29.945",
      "SH600547__buy_shares": "200",
      "SH600547__note": "底仓",
      "SH600547__sell_price": "",
      "SH600547__sell_date": "",
      "SZ300347__buy_price": "42.035",
      ...
      "__notes": "今日感想 ...",
      "__exported_at": "2026-05-26T..."
    }

行为:
  - Positions: 只 update 现存 row 的 5 个手填列 (buy_price/buy_shares/note/
    sell_price/sell_date), 不新建行. Symbol 找不到 → 跳过 (model 已 rotation
    出去, 不应再 ghost-update).
  - Notes: 追加/重写 USER_NOTES_MARKER 之后的 col-0 行. 不动 NOTES_TEXT 固定参考
    psychology 段.
  - 字段空值 ("") → 不写入 (避免覆盖 server 端通过 paper_trade 自动落值).
  - 类型: buy_price/sell_price → float, buy_shares → int, sell_date → str
    (YYYY-MM-DD, 不转 datetime — 跟 portfolio_excel 原始落值习惯一致).

Idempotent: 多次跑同 user_input.json 结果一致.

Exit codes:
  0 = ok (含 "no user_input.json, nothing to merge")
  1 = read/parse error (xlsx 损坏 / json 损坏) — 仅在 main() CLI 模式打印 + exit;
      被 portfolio_excel.py try-call 时只 print warning, 不抛.

Run:
    python examples/merge_user_input.py
    python examples/merge_user_input.py --dry-run   # 只 print 不写
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
USER_JSON = ROOT / "data_cache" / "user_input.json"
XLSX = ROOT / "data_cache" / "portfolio.xlsx"

# Positions sheet 列号 — 必须跟 portfolio_excel.py POS_COL_* 保持同步.
# (这里硬编码而不 import portfolio_excel 是为了避免循环依赖, schema 变了得 grep 同步).
POS_COL_SYMBOL = 2
POS_COL_STATUS = 17

FIELD_COL: dict[str, int] = {
    "buy_price": 10,    # 实际买入价
    "buy_shares": 11,   # 实际买入数
    "sell_price": 18,   # 卖出价
    "sell_date": 19,    # 卖出日期
    "note": 21,         # 备注
}

ALLOWED_FIELDS = set(FIELD_COL.keys())

NOTES_SHEET = "Notes"
USER_NOTES_MARKER = "═══ User Notes (from dashboard) ═══"


def _coerce_value(field: str, raw: Any) -> Any | None:
    """字符串/数字 → 强类型. 不可解析或空 → None (= 不 update)."""
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip() == "":
        return None
    if field in ("buy_price", "sell_price"):
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None
    if field == "buy_shares":
        try:
            return int(float(raw))
        except (TypeError, ValueError):
            return None
    if field == "sell_date":
        s = str(raw).strip()
        if not s:
            return None
        # 只取 YYYY-MM-DD 前 10 字符 (浏览器 <input type=date> 是 YYYY-MM-DD)
        return s[:10]
    if field == "note":
        s = str(raw).strip()
        return s if s else None
    return None


def _split_key(key: str) -> tuple[str, str] | None:
    """'SH600547__buy_price' → ('SH600547', 'buy_price'); 反之 None."""
    if not isinstance(key, str) or "__" not in key:
        return None
    sym, field = key.rsplit("__", 1)
    if not sym or field not in ALLOWED_FIELDS:
        return None
    return sym, field


def merge_positions(wb, data: dict, *, dry_run: bool = False) -> dict[str, int]:
    """Update Positions rows. Returns counts dict."""
    sh = wb["Positions"]
    # 建 sym → row index
    sym_to_row: dict[str, int] = {}
    for r in range(2, sh.max_row + 1):
        sym = sh.cell(r, POS_COL_SYMBOL).value
        if sym:
            sym_to_row[str(sym)] = r

    counts = {"updated": 0, "skipped_no_row": 0, "skipped_bad_value": 0, "skipped_meta": 0}
    for key, raw in data.items():
        if key.startswith("__"):
            counts["skipped_meta"] += 1
            continue
        parsed = _split_key(key)
        if parsed is None:
            counts["skipped_bad_value"] += 1
            continue
        sym, field = parsed
        row = sym_to_row.get(sym)
        if row is None:
            counts["skipped_no_row"] += 1
            continue
        value = _coerce_value(field, raw)
        if value is None:
            counts["skipped_bad_value"] += 1
            continue
        if not dry_run:
            sh.cell(row, FIELD_COL[field], value)
        counts["updated"] += 1
    return counts


def merge_notes(wb, notes_text: str | None, *, dry_run: bool = False) -> int:
    """Replace user-notes section in Notes sheet (after USER_NOTES_MARKER row).

    Returns number of lines written. notes_text=None or '' → noop (returns 0).
    保留 NOTES_TEXT 固定参考段 (marker 之前的 row).
    """
    if not notes_text:
        return 0
    if NOTES_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(NOTES_SHEET)
    else:
        ws = wb[NOTES_SHEET]

    # 找 marker row
    marker_row = None
    last_row = ws.max_row
    for r in range(1, last_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and USER_NOTES_MARKER in v:
            marker_row = r
            break

    lines = notes_text.split("\n")

    if dry_run:
        return len(lines)

    # 删除 marker 之后所有 row (col 0); 没 marker 则在 last_row+2 加 marker
    if marker_row is not None:
        # delete rows after marker
        if ws.max_row > marker_row:
            ws.delete_rows(marker_row + 1, ws.max_row - marker_row)
        start_row = marker_row + 1
    else:
        marker_row = (ws.max_row or 0) + 2  # 空一行后追加
        cell = ws.cell(marker_row, 1, USER_NOTES_MARKER)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor="305496")
        start_row = marker_row + 1

    for i, line in enumerate(lines):
        ws.cell(start_row + i, 1, line)
    return len(lines)


def run_merge(
    *,
    user_json_path: Path = USER_JSON,
    xlsx_path: Path = XLSX,
    dry_run: bool = False,
    verbose: bool = True,
) -> dict[str, Any]:
    """Programmatic entry — return summary dict. Idempotent.

    Returns:
      {"ok": bool, "reason": str|None, "positions": {...}, "notes_lines": int}
    """
    summary: dict[str, Any] = {
        "ok": True, "reason": None,
        "positions": {"updated": 0, "skipped_no_row": 0,
                       "skipped_bad_value": 0, "skipped_meta": 0},
        "notes_lines": 0,
        "dry_run": dry_run,
    }
    if not user_json_path.exists():
        summary["reason"] = f"no {user_json_path.name}, nothing to merge"
        if verbose:
            print(f"[merge_user_input] {summary['reason']}")
        return summary

    if not xlsx_path.exists():
        summary["ok"] = False
        summary["reason"] = f"portfolio.xlsx missing at {xlsx_path}"
        if verbose:
            print(f"[merge_user_input] WARN {summary['reason']}")
        return summary

    try:
        data = json.loads(user_json_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as e:
        summary["ok"] = False
        summary["reason"] = f"{type(e).__name__}: {e}"
        if verbose:
            print(f"[merge_user_input] WARN parse failed: {summary['reason']}")
        return summary

    if not isinstance(data, dict):
        summary["ok"] = False
        summary["reason"] = f"user_input.json top-level must be object, got {type(data).__name__}"
        if verbose:
            print(f"[merge_user_input] WARN {summary['reason']}")
        return summary

    notes_text = data.get("__notes")
    if notes_text is not None and not isinstance(notes_text, str):
        notes_text = str(notes_text)

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:  # noqa: BLE001
        summary["ok"] = False
        summary["reason"] = f"openpyxl load failed: {type(e).__name__}: {e}"
        if verbose:
            print(f"[merge_user_input] WARN {summary['reason']}")
        return summary

    pos_counts = merge_positions(wb, data, dry_run=dry_run)
    notes_lines = merge_notes(wb, notes_text, dry_run=dry_run)
    summary["positions"] = pos_counts
    summary["notes_lines"] = notes_lines

    if not dry_run:
        try:
            wb.save(xlsx_path)
        except Exception as e:  # noqa: BLE001
            summary["ok"] = False
            summary["reason"] = f"openpyxl save failed: {type(e).__name__}: {e}"
            if verbose:
                print(f"[merge_user_input] WARN {summary['reason']}")
            return summary

    if verbose:
        prefix = "[merge_user_input dry-run]" if dry_run else "[merge_user_input]"
        print(
            f"{prefix} updated={pos_counts['updated']} fields  "
            f"notes_lines={notes_lines}  "
            f"skipped(no_row={pos_counts['skipped_no_row']}, "
            f"bad_value={pos_counts['skipped_bad_value']}, "
            f"meta={pos_counts['skipped_meta']})"
        )
    return summary


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Merge user_input.json → portfolio.xlsx")
    p.add_argument("--user-json", default=str(USER_JSON))
    p.add_argument("--xlsx", default=str(XLSX))
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args(argv)
    summary = run_merge(
        user_json_path=Path(args.user_json),
        xlsx_path=Path(args.xlsx),
        dry_run=args.dry_run,
        verbose=True,
    )
    return 0 if summary["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())
