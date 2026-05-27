"""自动维护 data_cache/portfolio.xlsx — 4 sheet 实盘跟踪.

Sheet:
  Positions  持仓明细 (今日 BUY 追加 + 当前价每日更新, 公式自动算浮盈)
  Daily      每日总览 (跨日 P&L)
  Weekly     周结 (用户手填周区间/周收益, 模板)
  Training   月度训练 log (用户手填 val_loss/备注, 模板)

依赖:
  data_cache/paper_trade_log.csv  (今日 BUY/SELL 信号)
  data_cache/portfolio_state.json (当前持仓 symbols)
  data_cache/baidu_kline.parquet  (latest close)
  data_cache/universe.csv         (名称)

资金分配: score 加权
  total_capital = 50000
  daily_pool = total_capital / (K / N_DROP) = 50000 / 4 = 12500
  pick_quota = daily_pool * (score_i / sum_today_scores)
  shares = floor(pick_quota / price / 100) * 100

Run:  python examples/portfolio_excel.py
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data_cache" / "portfolio.xlsx"
STATE = ROOT / "data_cache" / "portfolio_state.json"
LOG = ROOT / "data_cache" / "paper_trade_log.csv"
KLINE = ROOT / "data_cache" / "baidu_kline.parquet"
UNIVERSE = ROOT / "data_cache" / "universe.csv"

TOTAL_CAPITAL = 50000.0  # 实盘本金 5 万 (backtest 内部 capital=25k 只是 alpha-集中参数, 不是本金)
K = 8
N_DROP = 2
DAILY_POOL = TOTAL_CAPITAL / (K / N_DROP)  # = 12500

POS_HEADERS = [
    "推荐日期", "代码", "名称", "Score", "推荐价",
    "Score权重%", "推荐金额", "推荐手数",
    "止损价(-8%)",  # col 9: 心理参考 (推荐价 ×0.92), 非机械触发 — backtest 证明机械止损反而扩大 MDD
    "实际买入价", "实际买入数", "实际成本",
    "当前价", "当前市值", "浮盈%", "浮盈元",
    "状态", "卖出价", "卖出日期", "实现盈亏", "备注",
]

# Positions sheet column indices (1-based, after 止损价 moved from col 21 → col 9)
POS_COL_DATE = 1
POS_COL_SYMBOL = 2
POS_COL_NAME = 3
POS_COL_SCORE = 4
POS_COL_REC_PRICE = 5
POS_COL_WEIGHT_PCT = 6
POS_COL_REC_AMOUNT = 7
POS_COL_REC_SHARES = 8
POS_COL_STOP_LOSS = 9       # was col 21
POS_COL_BUY_PRICE = 10      # was col 9
POS_COL_BUY_SHARES = 11     # was col 10
POS_COL_COST = 12           # was col 11
POS_COL_CURRENT = 13        # was col 12
POS_COL_MARKET_VALUE = 14   # was col 13
POS_COL_PNL_PCT = 15        # was col 14
POS_COL_PNL_YUAN = 16       # was col 15
POS_COL_STATUS = 17         # was col 16
POS_COL_SELL_PRICE = 18     # was col 17
POS_COL_SELL_DATE = 19      # was col 18
POS_COL_REALIZED_PNL = 20   # was col 19
POS_COL_NOTE = 21           # was col 20
DAILY_HEADERS = ["日期", "总资产", "现金", "持仓市值",
                  "当日P&L", "累计P&L", "收益率%", "持仓数"]
WEEKLY_HEADERS = ["周次(YYYY-WW)", "周初资产", "周末资产",
                   "周P&L", "周收益%", "交易次数", "胜率%"]
TRAINING_HEADERS = ["训练日期", "训练起", "训练止",
                     "样本数", "val_loss", "best_iter", "备注"]

# ---------- Forward OOS Track ----------
# 真实 forward OOS 始于 2026-05-25 (production train24 v19.1 部署日 2026-05-24 之后第一个交易日).
# 这里累积月度组合表现, 12 月后给出第一份完全干净 (无 backtest leak / 无 meta-leak)
# 的 forward OOS Sharpe / Calmar / MDD. 在此之前 backtest 60m Sharpe 1.09 / 24m OOS Sharpe 0.87
# 都仅供参考.

FORWARD_OOS_START = "2026-05-25"
BACKTEST_REF_SHARPE = 1.09  # production train24 60m backtest (含 meta-leak)
BACKTEST_OOS_24M_SHARPE = 0.87  # 24m strict OOS

FORWARD_OOS_HEADERS = [
    "month", "month_start_date", "month_end_date",
    "start_value", "end_value", "monthly_return",
    "cum_return", "picks_taken", "actual_buys", "n_picks_used",
    "notes",
]

# Alert level 染色 (forward_oos_monitor.py 注入)
ALERTS_CSV_PATH = "data_cache/forward_oos_alerts.csv"
ALERT_LEVEL_FILL = {
    "green": "C6EFCE",   # 浅绿
    "yellow": "FFEB9C",  # 浅黄
    "orange": "FFC299",  # 浅橙
    "red": "FFC7CE",     # 浅红
    "black": "808080",   # 深灰
}
FORWARD_OOS_STATS_HEADERS = ["指标", "值"]

MIN_MONTHS_FOR_ANN = 3
MIN_MONTHS_FOR_SHARPE = 6
MIN_MONTHS_FOR_MDD = 3


def _sym_to_code(sym: str) -> str:
    """SH600150 -> 600150 / SZ300661 -> 300661."""
    if isinstance(sym, str) and len(sym) >= 8 and sym[:2] in ("SH", "SZ"):
        return sym[2:]
    return str(sym).zfill(6)


def _month_end_value(
    kline_df: pd.DataFrame,
    picks: list[str],
    month: pd.Period,
    start_value: float,
) -> float | None:
    """估算月末组合总价值 (简化等权口径).

    - 月初按 picks 等权分 start_value
    - 每只 stock 用月内首交易日 close 作 entry, 月内最后一交易日 close 作 exit
    - 找不到价 → 该股贡献 0 收益
    - 全部找不到价 → 视作整月空仓返回 start_value
    - 空 picks → 整月空仓返回 start_value (现金不变)
    """
    if not picks:
        return start_value
    if kline_df is None or kline_df.empty:
        return None
    per_stock = start_value / len(picks)
    end_total = 0.0
    found_any = False
    for sym in picks:
        code = _sym_to_code(sym)
        sub = kline_df[kline_df["code"] == code]
        if sub.empty:
            end_total += per_stock
            continue
        period = sub["date"].dt.to_period("M")
        sub = sub[period == month].sort_values("date")
        if sub.empty:
            end_total += per_stock
            continue
        entry = float(sub.iloc[0]["close"])
        exit_ = float(sub.iloc[-1]["close"])
        if entry <= 0:
            end_total += per_stock
            continue
        ret = (exit_ - entry) / entry
        end_total += per_stock * (1.0 + ret)
        found_any = True
    if not found_any:
        return start_value
    return end_total


def compute_forward_oos_track(
    log_df: pd.DataFrame,
    kline_df: pd.DataFrame,
    holdings: list[str],
    total_capital: float = TOTAL_CAPITAL,
    start_date: str = FORWARD_OOS_START,
) -> pd.DataFrame:
    """累积 forward OOS 月度表现 DataFrame.

    Pure function (no I/O):
      log_df: [date, action, symbol, name, score, price]
      kline_df: [code (6 位 str), date (datetime64), close]
      holdings: 当前持仓 symbol 列表 (SH/SZ + 6 位)
      total_capital: 月初 base value (default 50000)
      start_date: forward OOS 起始日 (default 2026-05-25)

    Output DataFrame 列 = FORWARD_OOS_HEADERS, 每行一个月.
    空 log → 返回空 DataFrame (列齐, 0 行, 不崩).
    """
    cols = FORWARD_OOS_HEADERS
    if log_df is None or log_df.empty:
        return pd.DataFrame(columns=cols)

    df = log_df.copy()
    df["date"] = pd.to_datetime(df["date"])
    start_ts = pd.Timestamp(start_date)
    df = df[df["date"] >= start_ts]
    if df.empty:
        return pd.DataFrame(columns=cols)

    df["month"] = df["date"].dt.to_period("M")
    months = sorted(df["month"].unique())

    rows: list[dict] = []
    start_value = total_capital
    cum_factor = 1.0

    for month in months:
        m_df = df[df["month"] == month]
        m_start = m_df["date"].min()
        m_end = m_df["date"].max()

        buys = m_df[m_df["action"] == "BUY"]
        picks_taken = sorted(buys["symbol"].dropna().unique().tolist())

        is_last = (month == months[-1])
        actual_buys = sorted(holdings) if is_last else picks_taken
        n_picks_used = len(picks_taken)

        end_value = _month_end_value(kline_df, picks_taken, month, start_value)
        if end_value is None or start_value <= 0:
            monthly_return = float("nan")
            cum_return = float("nan")
        else:
            monthly_return = (end_value - start_value) / start_value
            cum_factor *= (1.0 + monthly_return)
            cum_return = cum_factor - 1.0

        def _r(x: float, n: int) -> float | None:
            return round(x, n) if x == x else None  # NaN check

        rows.append({
            "month": str(month),
            "month_start_date": m_start.date().isoformat(),
            "month_end_date": m_end.date().isoformat(),
            "start_value": round(start_value, 2),
            "end_value": (round(end_value, 2) if end_value is not None else None),
            "monthly_return": _r(monthly_return, 6),
            "cum_return": _r(cum_return, 6),
            "picks_taken": ",".join(picks_taken) if picks_taken else "",
            "actual_buys": ",".join(actual_buys) if actual_buys else "",
            "n_picks_used": n_picks_used,
            "notes": "",
        })

        if end_value is not None:
            start_value = end_value

    return pd.DataFrame(rows, columns=cols)


def compute_forward_oos_stats(
    track_df: pd.DataFrame,
    backtest_sharpe: float = BACKTEST_REF_SHARPE,
) -> dict[str, object]:
    """Forward OOS 累计统计.

    n<3: ann_return / MDD / Calmar = NaN
    n<6: Sharpe = NaN
    Sharpe = mean(monthly_ret) / std(monthly_ret) * sqrt(12) (rf=0).
    MDD: cum_curve 回撤最大值. Calmar = ann / |MDD|.
    """
    nan = float("nan")
    backtest_key = f"对比 backtest Sharpe {backtest_sharpe}"
    out: dict[str, object] = {
        "累积月数": 0,
        "forward cum return %": 0.0,
        "forward ann return %": nan,
        "forward Sharpe": nan,
        "forward MDD %": nan,
        "forward Calmar": nan,
        backtest_key: "等待 >=6 月样本",
        "验证里程碑": "12 月需达 (Sharpe>=0.5, Calmar>=0.8) 维持 production",
    }
    if track_df is None or track_df.empty:
        return out

    rets = pd.to_numeric(track_df["monthly_return"], errors="coerce").dropna()
    n = len(rets)
    out["累积月数"] = int(n)
    if n == 0:
        return out

    cum_factor = float((1.0 + rets).prod())
    cum_return = cum_factor - 1.0
    out["forward cum return %"] = round(cum_return * 100, 4)

    if n >= MIN_MONTHS_FOR_ANN:
        ann = (cum_factor ** (12.0 / n)) - 1.0
        out["forward ann return %"] = round(ann * 100, 4)

    if n >= MIN_MONTHS_FOR_SHARPE:
        mu = float(rets.mean())
        sigma = float(rets.std(ddof=1))
        if sigma > 0:
            sharpe = (mu / sigma) * (12 ** 0.5)
            out["forward Sharpe"] = round(sharpe, 4)
            out[backtest_key] = (
                f"forward={sharpe:.2f} vs backtest={backtest_sharpe:.2f} "
                f"(diff={sharpe - backtest_sharpe:+.2f})"
            )

    if n >= MIN_MONTHS_FOR_MDD:
        curve = (1.0 + rets).cumprod()
        peak = curve.cummax()
        dd = (curve - peak) / peak
        mdd = float(dd.min())
        out["forward MDD %"] = round(mdd * 100, 4)
        ann_pct = out["forward ann return %"]
        if mdd < 0 and isinstance(ann_pct, float) and ann_pct == ann_pct:
            calmar = (ann_pct / 100.0) / abs(mdd)
            out["forward Calmar"] = round(calmar, 4)
    return out


def _load_forward_inputs() -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Load (log_df, kline_df, holdings) — I/O wrapper around pure functions."""
    if LOG.exists():
        log_df = pd.read_csv(LOG, dtype={"symbol": str})
    else:
        log_df = pd.DataFrame(columns=["date", "action", "symbol", "name", "score", "price"])
    kline_df = pd.read_parquet(KLINE, columns=["code", "date", "close"])
    kline_df["code"] = kline_df["code"].astype(str).str.zfill(6)
    if STATE.exists():
        state = json.loads(STATE.read_text())
        holdings = list(state.get("holdings", []))
    else:
        holdings = []
    return log_df, kline_df, holdings


def _load_latest_alert_levels_by_month() -> dict[str, str]:
    """读 data_cache/forward_oos_alerts.csv, 按 month (YYYY-MM) 聚合, 返回最严重等级.
    优先级: black > red > orange > yellow > green.
    缺文件 → 返回空 dict (后续不染色).
    """
    path = Path(ALERTS_CSV_PATH)
    if not path.is_absolute():
        # 相对于 portfolio_excel.py 所在的 ROOT
        path = ROOT / ALERTS_CSV_PATH
    if not path.exists():
        return {}
    df = pd.read_csv(path, dtype={"date": str, "level": str})
    if df.empty:
        return {}
    df["month"] = df["date"].str[:7]  # 'YYYY-MM'
    priority = {"green": 0, "yellow": 1, "orange": 2, "red": 3, "black": 4}
    df["pri"] = df["level"].map(priority).fillna(0).astype(int)
    worst = df.sort_values(["month", "pri"]).groupby("month").tail(1)
    return dict(zip(worst["month"], worst["level"]))


def build_forward_oos_track_sheet(wb: Workbook) -> tuple[pd.DataFrame, dict[str, object]]:
    """生成/重建 "Forward OOS Track" + "Forward OOS Track Stats" 两个 sheet.
    每次全量重写 (行数小, 简洁), 不破坏其他 sheet.
    返回 (track_df, stats_dict) 供调用方打印.
    """
    log_df, kline_df, holdings = _load_forward_inputs()
    track_df = compute_forward_oos_track(log_df, kline_df, holdings)
    stats = compute_forward_oos_stats(track_df)

    # Forward OOS Track sheet
    name = "Forward OOS Track"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    head = ws.cell(row=1, column=1, value=f"Started: {FORWARD_OOS_START}")
    head.font = Font(bold=True, color="FFFFFF")
    head.fill = PatternFill("solid", fgColor="305496")
    ws.cell(row=1, column=2,
            value="真实 OOS 累积 — 12 月后给出第一份完全干净 forward Sharpe/Calmar")
    # 加 alert_level 列 (forward_oos_monitor.py 注入)
    headers_with_alert = list(FORWARD_OOS_HEADERS) + ["alert_level"]
    for col, h in enumerate(headers_with_alert, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    for col in range(1, len(headers_with_alert) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # 读 alerts.csv 算每月最严重 level
    alert_by_month = _load_latest_alert_levels_by_month()
    alert_col_idx = len(headers_with_alert)  # 最后一列

    for i, row in enumerate(track_df.itertuples(index=False), start=3):
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
        ws.cell(row=i, column=6).number_format = "0.00%"
        ws.cell(row=i, column=7).number_format = "0.00%"
        # alert_level 单元格 + 染色
        month_key = str(row.month) if hasattr(row, "month") else ""
        level = alert_by_month.get(month_key, "")
        alert_cell = ws.cell(row=i, column=alert_col_idx, value=level)
        if level in ALERT_LEVEL_FILL:
            alert_cell.fill = PatternFill("solid", fgColor=ALERT_LEVEL_FILL[level])
            if level == "black":
                alert_cell.font = Font(color="FFFFFF", bold=True)

    # Stats sheet
    stats_name = "Forward OOS Track Stats"
    if stats_name in wb.sheetnames:
        del wb[stats_name]
    ws2 = wb.create_sheet(stats_name)
    for col, h in enumerate(FORWARD_OOS_STATS_HEADERS, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 60
    for i, (k, v) in enumerate(stats.items(), start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v)
    milestone_row = ws2.max_row + 2
    ws2.cell(row=milestone_row, column=1, value="验证里程碑详细").font = Font(bold=True)
    milestones = [
        ("3 月", "看 cum return + MDD 是否符号正确 (alpha 是否在); 样本太小无 Sharpe"),
        ("6 月", "首次出 forward Sharpe; 若 < 0 → 立即停 production 回 v17 LGB legacy"),
        ("12 月", "首份干净 Sharpe/Calmar/MDD; Sharpe>=0.5 且 Calmar>=0.8 → 维持; 否则降级 capital"),
        ("backtest 参考",
         f"60m Sharpe={BACKTEST_REF_SHARPE} (含 meta-leak), 24m strict OOS Sharpe={BACKTEST_OOS_24M_SHARPE}"),
    ]
    for j, (k, v) in enumerate(milestones, start=milestone_row + 1):
        ws2.cell(row=j, column=1, value=k)
        ws2.cell(row=j, column=2, value=v)

    return track_df, stats


def name_map() -> dict[str, str]:
    df = pd.read_csv(UNIVERSE, dtype={"code": str})
    df["code"] = df["code"].astype(str).str.zfill(6)
    out = {}
    for _, r in df.iterrows():
        prefix = "SH" if r["code"].startswith(("6", "9")) else "SZ"
        out[f"{prefix}{r['code']}"] = r["name"]
    return out


def latest_closes() -> dict[str, tuple[pd.Timestamp, float]]:
    """旧实现:读 baidu_kline.parquet。Phase 1 swap 后 baidu_kline 是 hfq,
    数值跟实盘 qfq 不符。保留供 backtest/historical 用,Excel "当前价" 列改用
    latest_closes_qfq()。"""
    df = pd.read_parquet(KLINE, columns=["code", "date", "close"])
    df["code"] = df["code"].astype(str).str.zfill(6)
    df["sym"] = df["code"].apply(
        lambda c: ("SH" if c.startswith(("6", "9")) else "SZ") + c)
    latest = df.sort_values("date").groupby("sym").tail(1).set_index("sym")
    return {sym: (row["date"], row["close"]) for sym, row in latest.iterrows()}


def latest_closes_qfq(syms: list[str]) -> dict[str, tuple[pd.Timestamp, float]]:
    """实时拉腾讯 qt.gtimg.cn 当前价(qfq,符合实盘市场价)。

    Sandbox-OK endpoint(per CLAUDE.md)。批量一次性拉,几股 ~1s。
    交易时段返回 latest tick;非交易时段返回 last close。

    Returns: {sym: (timestamp, qfq_close_price)}
    """
    import requests
    if not syms:
        return {}
    # syms: 'SH600519' / 'SZ300347' → 腾讯格式 sh600519 / sz300347
    tx_syms = [s.lower() for s in syms]
    url = "http://qt.gtimg.cn/q=" + ",".join(tx_syms)
    out: dict[str, tuple[pd.Timestamp, float]] = {}
    try:
        r = requests.get(url, timeout=10)
        r.encoding = "gbk"  # 腾讯返回 gbk 编码
        for line in r.text.strip().split("\n"):
            # line: v_sh600519="1~贵州茅台~600519~1683.50~..."
            if "=" not in line:
                continue
            var, val = line.split("=", 1)
            var = var.strip()
            val = val.strip().rstrip(";").strip('"')
            if not val or "~" not in val:
                continue
            parts = val.split("~")
            if len(parts) < 31:
                continue
            sym = var.replace("v_", "").upper()  # sh600519 → SH600519
            try:
                price = float(parts[3])
                date_str = parts[30][:8]  # YYYYMMDD
                ts = pd.Timestamp(date_str)
                out[sym] = (ts, price)
            except (ValueError, IndexError):
                continue
    except Exception as e:
        print(f"  [qfq fetch warn] {type(e).__name__}: {e}")
    return out


NOTES_TEXT = [
    ["⚠️ 实盘心理 + 决策提醒", ""],
    ["", ""],
    ["几何累乘: 跌后回本所需涨幅", ""],
    ["跌幅", "回本所需涨幅"],
    ["-5%", "+5.3%"],
    ["-10%", "+11.1%"],
    ["-15%", "+17.6%"],
    ["-20%", "+25.0%"],
    ["-30%", "+42.9%"],
    ["-36% (v18 稳健 MDD)", "+56.3%"],
    ["-50%", "+100%"],
    ["", ""],
    ["v18 稳健 回测后验 (CSI300 / 60月 / capital=2.5万 backtest)", ""],
    ["累计回报", "+962.80%"],
    ["年化", "+60.43%"],
    ["Sharpe", "1.18"],
    ["最大回撤", "-35.89%"],
    ["月胜率", "56.67%"],
    ["⚠️ 真实期望 (扣 survivorship + 实摩擦)", "+30~37% 年化, MDD -40~-45%"],
    ["", ""],
    ["v17 LGB 安全回退 (CSI300 / 60月 / 5万)", ""],
    ["累计回报", "+244.39%"],
    ["年化", "+28.06%"],
    ["Sharpe", "0.94"],
    ["最大回撤", "-27.26%"],
    ["实盘期望", "+11~15% 年化, MDD -30~-35%"],
    ["", ""],
    ["实盘单月跌幅 → 操作档 (v18 阈值)", ""],
    ["跌幅档", "建议"],
    ["-5% 以内", "正常波动, 继续"],
    ["-5% ~ -10%", "暂停建仓, 已建仓继续持有"],
    ["-10% ~ -20%", "警惕, 考虑减仓 50%"],
    ["-20% ~ -36%", "暂停策略, 重新评估 alpha 是否失效"],
    ["超 -36%", "全部清仓, 等下次模型重训后再回来"],
    ["", ""],
    ["心理诚实", ""],
    ["大多数人在 -20% 时割肉", ""],
    ["回测 +54% ≠ 实盘 +54%, 行为偏差吃掉一半", ""],
    ["建议: 先 10-30% 仓位试水 1-3 个月", ""],
    ["", ""],
    ["买入规则: 集合竞价 9:25 看开盘价 vs 推荐价", ""],
    ["开盘 +5% 以上", "全跳, 等次日新信号"],
    ["开盘 +3% ~ +5%", "只买涨幅小的 1 只"],
    ["开盘 +1% ~ +3%", "挂 推荐价 +2% 等回调"],
    ["开盘 -1% ~ +1%", "直接挂限价 = 开盘价 (最佳入场)"],
    ["开盘 -1% ~ -3%", "直接买 (比推荐价便宜更好)"],
    ["开盘 -3% ~ -5%", "只半仓买单价低的 1 只"],
    ["开盘 -5% ~ -8%", "暂停, 14:30 跑新 signal 看 model 是否换股"],
    ["开盘 -8% 到跌停", "全跳, 等次日新信号"],
    ["", ""],
    ["跌停限价档 (主板±10% / 创业科创±20%)", ""],
    ["主板 (60_/00_)", "下跌 -10% 就跌停, 不能买"],
    ["创业板 (300)", "下跌 -20% 才跌停"],
    ["科创板 (688)", "下跌 -20% 才跌停"],
    ["", ""],
    ["持仓期跌幅 (已买入后)", ""],
    ["单日跌 -5% 以内", "正常波动, HOLD"],
    ["单日跌 -5% ~ -10%", "HOLD (T+1 也卖不了)"],
    ["次日低开 -3% 以上", "可考虑止损 (但 backtest 无止损规则)"],
    ["建议", "信任 model drop=2 节奏自动换仓"],
    ["", ""],
    ["⚠️ 止损价 是心理参考, 不机械执行!", ""],
    ["默认 = 推荐价 × 0.92 (-8% 警戒线)", ""],
    ["v18 C 实验已证: 机械 -10% 止损 MDD 反增 (-51% vs -50%)", ""],
    ["whip-saw 锁损 — 卖在底, 反弹回买", ""],
    ["正确用法", ""],
    ["跌破止损价 = 评估 触发器", "看是 model 错了 还是 noise"],
    ["model 还在 Top-K → HOLD (信号未变)", ""],
    ["model 跌出 Top-K → 下次跑 paper_trade 自动 SELL", ""],
    ["单股跌穿 -15%(深度回撤)", "可考虑人工 cut, 但要承认 model 失灵"],
    ["", ""],
    ["数据局限", ""],
    ["Survivorship bias", "用 2026 CSI300 反推, 真实少 5-15pp"],
    ["除权除息 (XD/DR)", "model 不感知, 看到 XD/DR 警惕"],
    ["回测 vs 实盘", "回测假设完美执行, 实际有滑点+情绪"],
]


def init_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for name, headers in [
        ("Positions", POS_HEADERS), ("Daily", DAILY_HEADERS),
        ("Weekly", WEEKLY_HEADERS), ("Training", TRAINING_HEADERS),
    ]:
        ws = wb.create_sheet(name)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 13

    ws = wb.create_sheet("Notes")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50
    for i, (a, b) in enumerate(NOTES_TEXT, 1):
        ca = ws.cell(row=i, column=1, value=a)
        cb = ws.cell(row=i, column=2, value=b)
        if a and not b and a.startswith(("⚠️", "几何", "v17", "实盘单月",
                                          "心理", "买入", "数据")):
            ca.font = Font(bold=True, color="FFFFFF", size=12)
            ca.fill = PatternFill("solid", fgColor="C00000")
        elif a in ("跌幅", "跌幅档"):
            ca.font = Font(bold=True)
            cb.font = Font(bold=True)
            ca.fill = PatternFill("solid", fgColor="D9E1F2")
            cb.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.freeze_panes = "A2"

    return wb


def existing_positions_keys(wb: Workbook) -> set[tuple[str, str]]:
    ws = wb["Positions"]
    out = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            out.add((str(row[0]), str(row[1])))
    return out


def existing_daily_dates(wb: Workbook) -> set[str]:
    ws = wb["Daily"]
    out = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            out.add(str(row[0]))
    return out


def add_positions_row(ws, today: str, sym: str, name: str, score: float,
                       price: float, weight_pct: float, quota: float,
                       shares: int, note: str = "") -> None:
    r = ws.max_row + 1
    ws.cell(row=r, column=POS_COL_DATE, value=today)
    ws.cell(row=r, column=POS_COL_SYMBOL, value=sym)
    ws.cell(row=r, column=POS_COL_NAME, value=name)
    ws.cell(row=r, column=POS_COL_SCORE, value=round(score, 4))
    ws.cell(row=r, column=POS_COL_REC_PRICE, value=round(price, 2))
    ws.cell(row=r, column=POS_COL_WEIGHT_PCT, value=round(weight_pct, 1))
    ws.cell(row=r, column=POS_COL_REC_AMOUNT, value=round(quota, 0))
    ws.cell(row=r, column=POS_COL_REC_SHARES, value=shares)
    # 止损价 = 推荐价 ×0.92 (默认), 实盘成交后用户可手改为 实际买入价 ×0.92
    ws.cell(row=r, column=POS_COL_STOP_LOSS, value=round(price * 0.92, 2))
    # Excel col letters (new layout): J=10 实买价 K=11 实买数 L=12 实际成本 M=13 当前价
    # N=14 市值 O=15 浮盈% P=16 浮盈元 Q=17 状态 R=18 卖出价
    ws.cell(row=r, column=POS_COL_COST, value=f"=J{r}*K{r}")
    ws.cell(row=r, column=POS_COL_MARKET_VALUE, value=f"=M{r}*K{r}")
    ws.cell(row=r, column=POS_COL_PNL_PCT, value=f'=IF(J{r}>0,(M{r}-J{r})/J{r},"")')
    ws.cell(row=r, column=POS_COL_PNL_YUAN, value=f'=IF(J{r}>0,N{r}-L{r},"")')
    ws.cell(row=r, column=POS_COL_STATUS, value="推荐")  # 默认 model 推荐, 用户填 实买价后由 reconcile_status_from_buy_price() 自动 → "持仓"
    ws.cell(row=r, column=POS_COL_REALIZED_PNL,
            value=f'=IF(AND(R{r}>0,K{r}>0),(R{r}-J{r})*K{r},"")')
    if note:
        ws.cell(row=r, column=POS_COL_NOTE, value=note)
    ws.cell(row=r, column=POS_COL_PNL_PCT).number_format = "0.00%"


def reconcile_status_from_buy_price(ws) -> tuple[int, int]:
    """auto-upgrade '推荐' → '持仓' 当 实买价 + 实买数 都填了 (用户手动买入后).
    auto-downgrade '持仓' → '推荐' 当 实买价 被清空 (e.g. 错填回滚).
    返回 (upgraded, downgraded)."""
    upgraded = 0
    downgraded = 0
    for r in range(2, ws.max_row + 1):
        sym = ws.cell(row=r, column=POS_COL_SYMBOL).value
        if not sym:
            continue
        status = ws.cell(row=r, column=POS_COL_STATUS).value
        buy_price = ws.cell(row=r, column=POS_COL_BUY_PRICE).value
        buy_shares = ws.cell(row=r, column=POS_COL_BUY_SHARES).value
        has_real_buy = buy_price is not None and buy_shares is not None
        if has_real_buy and status == "推荐":
            ws.cell(row=r, column=POS_COL_STATUS, value="持仓")
            # 持仓状态绝不应有 sell_*; 清掉之前误标 '已平' 时遗留的 sell_price/sell_date/实现盈亏
            # NOTE: openpyxl ws.cell(r,c,None) 不真清值; 必须 .value = None 显式赋值
            ws.cell(row=r, column=POS_COL_SELL_PRICE).value = None
            ws.cell(row=r, column=POS_COL_SELL_DATE).value = None
            ws.cell(row=r, column=POS_COL_REALIZED_PNL).value = None
            upgraded += 1
        elif not has_real_buy and status == "持仓":
            ws.cell(row=r, column=POS_COL_STATUS, value="推荐")
            # 推荐状态同样不应有 sell_*
            ws.cell(row=r, column=POS_COL_SELL_PRICE).value = None
            ws.cell(row=r, column=POS_COL_SELL_DATE).value = None
            ws.cell(row=r, column=POS_COL_REALIZED_PNL).value = None
            downgraded += 1
    return upgraded, downgraded


def drop_stale_unfilled_recs(ws, today_str: str, holdings_set: set[str]) -> int:
    """删除 stale 推荐: 实际买入数为空 AND 推荐日期 < today AND sym NOT in holdings.
    holdings_set 是 portfolio_state.json 的 active model picks — 这些仍是模型 active 推荐,
    即使用户没买入也不删 (反复推荐 rotation 的常态).
    只删 模型已 SELL 出 holdings + 用户也没买入 的真正 stale 行.
    保留今日 (推荐日期 == today) 未买入信号 — 等当日实际执行.
    返回删除行数."""
    to_delete: list[int] = []
    for r in range(2, ws.max_row + 1):
        sym = ws.cell(row=r, column=POS_COL_SYMBOL).value
        if not sym:
            continue
        if sym in holdings_set:
            continue  # model 仍 active, 跳过
        rec_date = ws.cell(row=r, column=POS_COL_DATE).value
        if rec_date is None:
            continue
        rec_date_str = str(rec_date)[:10]
        actual_shares = ws.cell(row=r, column=POS_COL_BUY_SHARES).value
        if not actual_shares and rec_date_str < today_str:
            to_delete.append(r)
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)
    return len(to_delete)


def reconcile_with_state(ws, holdings_set: set[str], today_str: str) -> tuple[int, int]:
    """同步 Positions 跟 portfolio_state.
    任何不在 holdings 的行 (不论之前状态):
      - 实际买入数 为空 → 直接删 (从未真买, 不留垃圾)
      - 否则 → 标 '已平' + 填卖出日期 (真历史 SELL, 保留)
    返回 (deleted, closed)."""
    to_delete: list[int] = []
    closed = 0
    for r in range(2, ws.max_row + 1):
        sym = ws.cell(row=r, column=POS_COL_SYMBOL).value
        if not sym:
            continue
        if sym in holdings_set:
            continue  # 还在持仓里, 不处理
        actual_shares = ws.cell(row=r, column=POS_COL_BUY_SHARES).value
        if not actual_shares:
            to_delete.append(r)
        elif ws.cell(row=r, column=POS_COL_STATUS).value != "已平":
            ws.cell(row=r, column=POS_COL_STATUS, value="已平")
            ws.cell(row=r, column=POS_COL_SELL_DATE, value=today_str)
            closed += 1
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)
    return len(to_delete), closed


def update_current_prices(ws, closes: dict[str, tuple[pd.Timestamp, float]]) -> None:
    """每个 持仓 row 更新 col 12 当前价 + 重算 col 13 市值 / col 14 浮盈% / col 15 浮盈元
    (基于 col 9 实买价 + col 10 实买数 + col 12 当前价)."""
    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=POS_COL_STATUS).value
        if status != "持仓":
            continue
        sym = ws.cell(row=r, column=POS_COL_SYMBOL).value
        if sym in closes:
            _, px = closes[sym]
            cur = round(float(px), 2)
            ws.cell(row=r, column=POS_COL_CURRENT, value=cur)
            # 如有 实买价 + 实买数 → 重算 市值/浮盈%/浮盈元
            buy_price = ws.cell(row=r, column=POS_COL_BUY_PRICE).value
            buy_shares = ws.cell(row=r, column=POS_COL_BUY_SHARES).value
            if buy_price and buy_shares:
                bp = float(buy_price)
                bs = float(buy_shares)
                mv = round(cur * bs, 2)
                # 浮盈% 存为 fraction(Excel "0.00%" format 自动 ×100 显示)
                pnl_pct_frac = round((cur / bp - 1), 4) if bp else None
                pnl_yuan = round((cur - bp) * bs, 2)
                ws.cell(row=r, column=POS_COL_MARKET_VALUE, value=mv)
                ws.cell(row=r, column=POS_COL_PNL_PCT, value=pnl_pct_frac)
                ws.cell(row=r, column=POS_COL_PNL_PCT).number_format = "0.00%"
                ws.cell(row=r, column=POS_COL_PNL_YUAN, value=pnl_yuan)


def compute_daily_snapshot(ws_pos, closes) -> tuple[float, float, int]:
    """从 Positions sheet 算 持仓市值, 累计实际成本, 持仓数."""
    market_value = 0.0
    cost = 0.0
    n = 0
    for r in range(2, ws_pos.max_row + 1):
        status = ws_pos.cell(row=r, column=POS_COL_STATUS).value
        if status != "持仓":
            continue
        sym = ws_pos.cell(row=r, column=POS_COL_SYMBOL).value
        actual_price = ws_pos.cell(row=r, column=POS_COL_BUY_PRICE).value or 0
        actual_shares = ws_pos.cell(row=r, column=POS_COL_BUY_SHARES).value or 0
        if not actual_shares:
            continue
        cost += actual_price * actual_shares
        if sym in closes:
            _, px = closes[sym]
            market_value += px * actual_shares
        else:
            market_value += actual_price * actual_shares
        n += 1
    return market_value, cost, n


def add_daily_row(ws, today: str, market_value: float, cost: float,
                  n_holdings: int) -> None:
    cash = TOTAL_CAPITAL - cost
    total = cash + market_value
    pnl_cum = total - TOTAL_CAPITAL
    pnl_pct = pnl_cum / TOTAL_CAPITAL

    prev_total = TOTAL_CAPITAL
    if ws.max_row >= 2:
        prev = ws.cell(row=ws.max_row, column=2).value
        if isinstance(prev, (int, float)):
            prev_total = prev
    pnl_day = total - prev_total

    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=today)
    ws.cell(row=r, column=2, value=round(total, 2))
    ws.cell(row=r, column=3, value=round(cash, 2))
    ws.cell(row=r, column=4, value=round(market_value, 2))
    ws.cell(row=r, column=5, value=round(pnl_day, 2))
    ws.cell(row=r, column=6, value=round(pnl_cum, 2))
    ws.cell(row=r, column=7, value=round(pnl_pct, 4))
    ws.cell(row=r, column=8, value=n_holdings)
    ws.cell(row=r, column=7).number_format = "0.00%"


def main() -> None:
    if not LOG.exists():
        print(f"找不到 {LOG}, 先跑 paper_trade_today.py")
        return

    log_df = pd.read_csv(LOG, dtype={"symbol": str})
    state = json.loads(STATE.read_text())
    today = state.get("date", str(log_df["date"].max()))
    print(f"今日: {today}")

    today_buys = log_df[(log_df["date"] == today) & (log_df["action"] == "BUY")]
    if today_buys.empty:
        print(f"  今天 ({today}) 没有 BUY 信号 — 仅更新 Daily 总览")
    else:
        print(f"  今日 BUY 信号: {len(today_buys)} 只")
        print(today_buys[["symbol", "name", "score", "price"]].to_string(index=False))

    names = name_map()
    print(f"\n加载 universe 名称表: {len(names)} 只")
    print("加载 latest closes (parquet) ...")
    closes = latest_closes()
    latest_date = max(d for d, _ in closes.values())
    print(f"  最新交易日 (parquet): {latest_date.date()}")

    if XLSX.exists():
        print(f"\n打开 existing {XLSX.name}")
        wb = load_workbook(XLSX)
    else:
        print(f"\n创建新 {XLSX.name}")
        wb = init_workbook()

    ws_pos = wb["Positions"]
    pos_keys = existing_positions_keys(wb)

    if not today_buys.empty:
        # qfq 价 batch fetch — paper_trade_log price 是 hfq, Excel 显示统一 qfq
        today_qfq = latest_closes_qfq(today_buys["symbol"].dropna().unique().tolist())
        sum_score = today_buys["score"].sum()
        for _, row in today_buys.iterrows():
            sym = row["symbol"]
            key = (today, sym)
            if key in pos_keys:
                print(f"  跳过 {sym} (Positions 已有 {today})")
                continue
            weight = row["score"] / sum_score if sum_score > 0 else 0.5
            weight_pct = weight * 100
            quota = DAILY_POOL * weight
            # 优先 qfq (实盘对齐), fallback 到 hfq from log
            price = float(today_qfq[sym][1]) if sym in today_qfq else float(row["price"])
            shares = int(quota // (price * 100)) * 100 if price > 0 else 0
            note = ""
            if shares == 0 and price > 0 and price * 100 <= DAILY_POOL:
                shares = 100
                quota = price * 100
                note = f"score quota买不起1手, 已上调至1手({quota:.0f}元)"
            elif shares == 0 and price * 100 > DAILY_POOL:
                note = f"1手价{price*100:.0f}元 > daily_pool{DAILY_POOL:.0f} - 跳过"
            add_positions_row(
                ws_pos, today, sym, names.get(sym, "?"),
                row["score"], price, weight_pct, quota, shares, note,
            )
            print(f"  + Positions: {sym}  weight={weight_pct:.1f}%  "
                  f"quota={quota:.0f}  推荐手数={shares}  {note}")

    # auto-clean stale 推荐 (实买为空 + 推荐日期 < today + NOT in holdings)
    current_holdings = set(state.get("holdings", []))
    stale_dropped = drop_stale_unfilled_recs(ws_pos, today, current_holdings)
    if stale_dropped:
        print(f"\n[stale-clean] 删除 {stale_dropped} 行 (推荐日期早于今日, 未实际买入, 已 SELL 出 holdings)")

    # auto-reconcile 状态: 实买价填了 → '持仓', 否则 → '推荐'
    up, down = reconcile_status_from_buy_price(ws_pos)
    if up or down:
        msg = []
        if up: msg.append(f"{up} 行 推荐 → 持仓")
        if down: msg.append(f"{down} 行 持仓 → 推荐")
        print(f"[status-reconcile] " + " + ".join(msg))

    # 同步 Positions 跟 portfolio_state
    deleted, closed = reconcile_with_state(ws_pos, current_holdings, today)
    if deleted or closed:
        msg = []
        if deleted:
            msg.append(f"删除 {deleted} 行 (未成交推荐)")
        if closed:
            msg.append(f"标记 {closed} 行 '已平' (历史 SELL)")
        print(f"\n[reconcile] " + " + ".join(msg))

    # v19.4 fix: baidu_kline 是 hfq,当前价应用 qfq 实时报价。提取持仓股 → 拉腾讯 qfq → merge。
    holding_syms = set()
    for r in range(2, ws_pos.max_row + 1):
        if ws_pos.cell(row=r, column=POS_COL_STATUS).value == "持仓":
            sym = ws_pos.cell(row=r, column=POS_COL_SYMBOL).value
            if sym:
                holding_syms.add(sym)
    if holding_syms:
        qfq_closes = latest_closes_qfq(list(holding_syms))
        print(f"  qfq 实时报价(腾讯): {len(qfq_closes)}/{len(holding_syms)} 持仓股")
        for sym, (ts, px) in qfq_closes.items():
            closes[sym] = (ts, px)  # qfq override hfq

    update_current_prices(ws_pos, closes)

    daily_dates = existing_daily_dates(wb)
    if today in daily_dates:
        print(f"\nDaily ({today}) 已存在 — 跳过追加")
    else:
        mv, cost, n_h = compute_daily_snapshot(ws_pos, closes)
        add_daily_row(wb["Daily"], today, mv, cost, n_h)
        print(f"\n+ Daily {today}: 总资产={TOTAL_CAPITAL - cost + mv:.0f} "
              f"持仓市值={mv:.0f} 现金={TOTAL_CAPITAL - cost:.0f} 持仓数={n_h}")

    # Forward OOS Track (累积月度真实表现, 从 2026-05-25 起)
    track_df, stats = build_forward_oos_track_sheet(wb)
    if track_df.empty:
        print(f"\n[Forward OOS] 0 月样本 (forward OOS 起始日 {FORWARD_OOS_START} 之后无数据)")
    else:
        print(f"\n[Forward OOS] 累积 {len(track_df)} 月 / cum={stats['forward cum return %']}% "
              f"/ Sharpe={stats['forward Sharpe']}")

    wb.save(XLSX)
    print(f"\n保存: {XLSX} ({XLSX.stat().st_size/1024:.1f} KB)")

    # Phase 1 merge: 若 user 通过 dashboard 导出了 data_cache/user_input.json,
    # 自动 merge 到 Positions 手填字段 + Notes sheet. 失败不阻塞 (just warn).
    try:
        import importlib.util as _il
        _spec = _il.spec_from_file_location(
            "_mui", Path(__file__).resolve().parent / "merge_user_input.py",
        )
        if _spec and _spec.loader:
            _mod = _il.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            _mod.run_merge(verbose=True)
    except Exception as _e:  # noqa: BLE001
        print(f"[merge_user_input] WARN auto-merge failed: {type(_e).__name__}: {_e}")


if __name__ == "__main__":
    main()
