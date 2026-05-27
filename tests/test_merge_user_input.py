"""merge_user_input.py — unit tests (idempotent + boundary + edge cases).

合成 portfolio.xlsx + user_input.json → run_merge → 验证 openpyxl 读回字段对.
"""
from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

# load merge_user_input from examples/
_SPEC = importlib.util.spec_from_file_location(
    "merge_user_input_mod",
    Path(__file__).resolve().parent.parent / "examples" / "merge_user_input.py",
)
assert _SPEC and _SPEC.loader
mui = importlib.util.module_from_spec(_SPEC)
_SPEC.loader.exec_module(mui)


# ---------- fixtures ----------
def _make_xlsx(tmp_path: Path, symbols: list[str]) -> Path:
    """Build a synthetic portfolio.xlsx with given symbols in Positions sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Positions")
    headers = [
        "推荐日期", "代码", "名称", "Score", "推荐价", "Score权重%",
        "推荐金额", "推荐手数", "止损价(-8%)",
        "实际买入价", "实际买入数", "实际成本",
        "当前价", "当前市值", "浮盈%", "浮盈元",
        "状态", "卖出价", "卖出日期", "实现盈亏", "备注",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, sym in enumerate(symbols, start=2):
        ws.cell(i, 1, "2026-05-26")
        ws.cell(i, 2, sym)
        ws.cell(i, 3, f"NAME-{i}")
        ws.cell(i, 17, "推荐")
    # also create empty Notes sheet
    wb.create_sheet("Notes")
    wb["Notes"].cell(1, 1, "⚠️ existing reference text")
    wb["Notes"].cell(2, 1, "几何累乘: 跌后回本所需涨幅")
    path = tmp_path / "portfolio.xlsx"
    wb.save(path)
    return path


def _write_json(tmp_path: Path, data: dict) -> Path:
    path = tmp_path / "user_input.json"
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    return path


# ---------- 1. missing user_input.json ----------
def test_missing_user_json_is_ok(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    summary = mui.run_merge(
        user_json_path=tmp_path / "nope.json",
        xlsx_path=xlsx, verbose=False,
    )
    assert summary["ok"] is True
    assert "nothing to merge" in summary["reason"]


# ---------- 2. happy path Positions update ----------
def test_positions_update_happy_path(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547", "SZ300347"])
    js = _write_json(tmp_path, {
        "SH600547__buy_price": "29.945",
        "SH600547__buy_shares": "200",
        "SH600547__note": "底仓",
        "SZ300347__buy_price": "42.04",
        "SZ300347__buy_shares": "100",
        "__exported_at": "2026-05-26T00:00:00Z",
    })
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["ok"] is True
    assert summary["positions"]["updated"] == 5
    assert summary["positions"]["skipped_meta"] == 1

    wb = load_workbook(xlsx)
    ws = wb["Positions"]
    assert ws.cell(2, 10).value == pytest.approx(29.945)
    assert ws.cell(2, 11).value == 200
    assert ws.cell(2, 21).value == "底仓"
    assert ws.cell(3, 10).value == pytest.approx(42.04)
    assert ws.cell(3, 11).value == 100


# ---------- 3. unknown symbol skipped ----------
def test_unknown_symbol_skipped(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {
        "SH600547__buy_price": "10.5",
        "SZ999999__buy_price": "20",
    })
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["positions"]["updated"] == 1
    assert summary["positions"]["skipped_no_row"] == 1


# ---------- 4. unknown field skipped ----------
def test_unknown_field_skipped(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {
        "SH600547__buy_price": "10.5",
        "SH600547__hacker_field": "evil",
    })
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["positions"]["updated"] == 1
    assert summary["positions"]["skipped_bad_value"] == 1


# ---------- 5. empty string skipped (don't overwrite) ----------
def test_empty_string_skipped(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    # pre-fill xlsx with an existing value
    wb = load_workbook(xlsx)
    wb["Positions"].cell(2, 10, 99.99)
    wb.save(xlsx)
    js = _write_json(tmp_path, {"SH600547__buy_price": ""})
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["positions"]["updated"] == 0
    assert summary["positions"]["skipped_bad_value"] == 1
    # 原值未被覆盖
    wb2 = load_workbook(xlsx)
    assert wb2["Positions"].cell(2, 10).value == pytest.approx(99.99)


# ---------- 6. invalid number string skipped ----------
def test_invalid_number_skipped(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {"SH600547__buy_price": "not_a_number"})
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["positions"]["updated"] == 0
    assert summary["positions"]["skipped_bad_value"] == 1


# ---------- 7. sell_date type str ----------
def test_sell_date_str(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {"SH600547__sell_date": "2026-05-26"})
    mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    wb = load_workbook(xlsx)
    assert wb["Positions"].cell(2, 19).value == "2026-05-26"


# ---------- 8. buy_shares coerced to int even from "100.0" ----------
def test_buy_shares_int_coercion(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {"SH600547__buy_shares": "100.0"})
    mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    wb = load_workbook(xlsx)
    val = wb["Positions"].cell(2, 11).value
    assert val == 100
    assert isinstance(val, int)


# ---------- 9. notes appended with marker ----------
def test_notes_appended(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {
        "__notes": "今日 SH600547 实买 200 @ 29.945\n感想: 黄金强势",
    })
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["notes_lines"] == 2

    wb = load_workbook(xlsx)
    ws = wb["Notes"]
    # 原 reference text 不动
    assert ws.cell(1, 1).value == "⚠️ existing reference text"
    # marker + 2 行 user notes 应该在最后
    marker_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and mui.USER_NOTES_MARKER in v:
            marker_row = r
            break
    assert marker_row is not None
    assert ws.cell(marker_row + 1, 1).value == "今日 SH600547 实买 200 @ 29.945"
    assert ws.cell(marker_row + 2, 1).value == "感想: 黄金强势"


# ---------- 10. idempotent re-merge ----------
def test_idempotent_remerge(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {
        "SH600547__buy_price": "29.945",
        "SH600547__buy_shares": "200",
        "__notes": "n1\nn2\nn3",
    })
    s1 = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    s2 = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert s1["positions"]["updated"] == s2["positions"]["updated"]
    assert s1["notes_lines"] == s2["notes_lines"] == 3

    wb = load_workbook(xlsx)
    ws = wb["Notes"]
    marker_count = sum(
        1 for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 1).value, str)
        and mui.USER_NOTES_MARKER in ws.cell(r, 1).value
    )
    assert marker_count == 1


# ---------- 11. dry-run doesn't write ----------
def test_dry_run_no_write(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {"SH600547__buy_price": "12.34"})
    summary = mui.run_merge(
        user_json_path=js, xlsx_path=xlsx, dry_run=True, verbose=False,
    )
    assert summary["positions"]["updated"] == 1
    wb = load_workbook(xlsx)
    assert wb["Positions"].cell(2, 10).value is None


# ---------- 12. corrupt JSON returns ok=False but doesn't raise ----------
def test_corrupt_json_returns_error(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    bad = tmp_path / "bad.json"
    bad.write_text("{not valid json", encoding="utf-8")
    summary = mui.run_merge(user_json_path=bad, xlsx_path=xlsx, verbose=False)
    assert summary["ok"] is False
    assert "JSONDecodeError" in summary["reason"]


# ---------- 13. notes overwrite not append (re-export replaces) ----------
def test_notes_overwrite_not_append(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js1 = _write_json(tmp_path, {"__notes": "old line 1\nold line 2"})
    mui.run_merge(user_json_path=js1, xlsx_path=xlsx, verbose=False)
    # second export with different content
    js2 = tmp_path / "user_input.json"
    js2.write_text(json.dumps({"__notes": "new"}), encoding="utf-8")
    mui.run_merge(user_json_path=js2, xlsx_path=xlsx, verbose=False)
    wb = load_workbook(xlsx)
    ws = wb["Notes"]
    marker_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and mui.USER_NOTES_MARKER in v:
            marker_row = r
            break
    assert marker_row is not None
    # marker + 1 行 "new" + 不应有 old line
    assert ws.cell(marker_row + 1, 1).value == "new"
    assert ws.cell(marker_row + 2, 1).value in (None, "")


# ---------- 14. all 5 editable fields exercised ----------
def test_all_fields(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = _write_json(tmp_path, {
        "SH600547__buy_price": "29.945",
        "SH600547__buy_shares": "200",
        "SH600547__sell_price": "31.50",
        "SH600547__sell_date": "2026-06-01",
        "SH600547__note": "全部字段",
    })
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["positions"]["updated"] == 5
    wb = load_workbook(xlsx)
    ws = wb["Positions"]
    assert ws.cell(2, 10).value == pytest.approx(29.945)
    assert ws.cell(2, 11).value == 200
    assert ws.cell(2, 18).value == pytest.approx(31.50)
    assert ws.cell(2, 19).value == "2026-06-01"
    assert ws.cell(2, 21).value == "全部字段"


# ---------- 15. missing xlsx returns ok=False ----------
def test_missing_xlsx_returns_error(tmp_path: Path):
    js = _write_json(tmp_path, {"SH600547__buy_price": "10"})
    summary = mui.run_merge(
        user_json_path=js,
        xlsx_path=tmp_path / "no_such.xlsx",
        verbose=False,
    )
    assert summary["ok"] is False
    assert "missing" in summary["reason"]


# ---------- 16. non-dict JSON returns error ----------
def test_non_dict_json_error(tmp_path: Path):
    xlsx = _make_xlsx(tmp_path, ["SH600547"])
    js = tmp_path / "list.json"
    js.write_text("[1,2,3]", encoding="utf-8")
    summary = mui.run_merge(user_json_path=js, xlsx_path=xlsx, verbose=False)
    assert summary["ok"] is False
    assert "must be object" in summary["reason"]
