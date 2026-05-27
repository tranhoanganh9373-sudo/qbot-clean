"""Holdings 行业集中度 — pie chart + HHI + 行业权重表.

数据源:
  - data_cache/portfolio_state.json    7 holdings (current)
  - data_cache/industry/industry_membership.parquet  SW level-1 (5203 行)
  - data_cache/portfolio.xlsx Positions sheet 实际买入价×买入数 (真持仓 2 股)

逻辑:
  1. 加载 holdings list (portfolio_state.holdings).
  2. Positions 有 (实际买入价 × 实际买入数) → 用实际市值; 缺则用等权.
  3. join industry_membership (SW level-1) → 按 industry_name 聚合.
  4. HHI = Σ(pct_i)² / 10000, range [0,1].
  5. Pie + bar matplotlib → base64 → HTML; HHI 着色;行业权重表着色.
"""
from __future__ import annotations

import html
import json
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook

from dashboard.utils.fig_to_base64 import fig_to_base64


def _ensure_cjk_font() -> None:
    """把系统 CJK 字体 prepend 到 sans-serif 链, 防中文渲染成方框.

    幂等; 与 factor_ic_heatmap._ensure_cjk_font 同样的候选链.
    """
    candidates = (
        "PingFang SC", "PingFang HK", "Heiti TC", "STHeiti",
        "Hiragino Sans GB", "Songti SC", "Arial Unicode MS",
        "Noto Sans CJK SC", "SimHei",
    )
    available = {f.name for f in fm.fontManager.ttflist}
    chain = [c for c in candidates if c in available]
    if not chain:
        return
    current = plt.rcParams["font.sans-serif"]
    merged = chain + [f for f in current if f not in chain]
    plt.rcParams["font.sans-serif"] = merged
    plt.rcParams["axes.unicode_minus"] = False


_ensure_cjk_font()

# ----- 常量 --------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent.parent
DC = ROOT / "data_cache"

STATE_PATH = DC / "portfolio_state.json"
INDUSTRY_PARQUET = DC / "industry" / "industry_membership.parquet"
PORTFOLIO_XLSX = DC / "portfolio.xlsx"

# Positions sheet col idx (1-based, openpyxl 风格); 由 portfolio.xlsx header 实测得来
POS_SHEET = "Positions"
POS_COL_CODE = 2          # '代码'
POS_COL_BUY_PRICE = 10    # '实际买入价'
POS_COL_BUY_SHARES = 11   # '实际买入数'
POS_COL_STATUS = 17       # '状态'

DEFAULT_EQUAL_WEIGHT_NOTIONAL = 50_000  # 等权 fallback 时的总池 ¥

# HHI 阈值 (经典分类)
HHI_LOW = 0.18      # < 低集中
HHI_MID = 0.25      # 0.18–0.25 中集中, > 高集中

# 单行业占比阈值
PCT_RED = 50.0      # > 50% 标红
PCT_ORANGE = 30.0   # > 30% 橙色


# ----- 辅助函数 ----------------------------------------------------------
def _placeholder(message: str) -> str:
    return f'<div class="placeholder-content">{message}</div>'


def _code_to_sym(code: str) -> str:
    """industry_membership.code (e.g. '600519') → 'SH600519' / 'SZ000001'."""
    if not code:
        return ""
    return ("SH" if code.startswith(("6", "9")) else "SZ") + code


def _load_actual_weights(holdings: list[str], xlsx_path: Path) -> dict[str, float]:
    """从 Positions sheet 读 (实际买入价 × 实际买入数) → 市值; 不存在则 {}."""
    weights: dict[str, float] = {}
    if not xlsx_path.exists():
        return weights
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return weights
    try:
        if POS_SHEET not in wb.sheetnames:
            return weights
        sh = wb[POS_SHEET]
        for r in range(2, sh.max_row + 1):
            sym = sh.cell(r, POS_COL_CODE).value
            if sym not in holdings:
                continue
            buy_price = sh.cell(r, POS_COL_BUY_PRICE).value
            buy_shares = sh.cell(r, POS_COL_BUY_SHARES).value
            if isinstance(buy_price, (int, float)) and isinstance(buy_shares, (int, float)):
                if buy_price > 0 and buy_shares > 0:
                    # 同一 sym 取最新一行 (覆盖)
                    weights[sym] = float(buy_price) * float(buy_shares)
    finally:
        wb.close()
    return weights


def _classify_hhi(hhi: float) -> tuple[str, str]:
    """返回 (色, 中文标签)."""
    if hhi < HHI_LOW:
        return ("#16a34a", "低集中")
    if hhi < HHI_MID:
        return ("#f59e0b", "中集中")
    return ("#dc2626", "高集中风险")


def _pct_color(pct: float) -> str:
    if pct > PCT_RED:
        return "#dc2626"
    if pct > PCT_ORANGE:
        return "#f59e0b"
    return "inherit"


# ----- 主入口 ------------------------------------------------------------
def build_sector_breakdown_section() -> str:
    """7 holdings × SW level-1 行业 pie + HHI + 表. 返回 HTML fragment."""
    # 1. holdings
    if not STATE_PATH.exists():
        return _placeholder(f"未找到 <code>{STATE_PATH.name}</code>")
    try:
        state = json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        return _placeholder(f"portfolio_state.json 解析失败: <code>{html.escape(str(exc))}</code>")
    holdings = list(state.get("holdings", []))
    if not holdings:
        return _placeholder("portfolio_state.json holdings 为空")

    # 2. industry membership
    if not INDUSTRY_PARQUET.exists():
        return _placeholder("未找到 <code>industry/industry_membership.parquet</code>")
    try:
        ind = pd.read_parquet(INDUSTRY_PARQUET)
    except Exception as exc:
        return _placeholder(f"industry parquet 读取失败: <code>{html.escape(str(exc))}</code>")

    needed_cols = {"code", "industry_name"}
    if not needed_cols.issubset(ind.columns):
        return _placeholder(
            f"industry parquet 缺列 {needed_cols - set(ind.columns)}"
        )
    ind = ind[["code", "industry_name"]].copy()
    ind["sym"] = ind["code"].astype(str).map(_code_to_sym)
    # 同一 sym 可能有多版 membership (历史); 取首条
    ind = ind.drop_duplicates(subset="sym", keep="first")

    # 3. weights: 实际市值优先, 缺则等权
    actual = _load_actual_weights(holdings, PORTFOLIO_XLSX)
    weights: dict[str, float] = {}
    equal_notional = DEFAULT_EQUAL_WEIGHT_NOTIONAL / len(holdings)
    n_actual = 0
    for sym in holdings:
        if sym in actual:
            weights[sym] = actual[sym]
            n_actual += 1
        else:
            weights[sym] = equal_notional
    weighting_basis = (
        f"{n_actual}/{len(holdings)} 持仓用 Positions 实际买入市值, 余 {len(holdings) - n_actual} 等权"
        if n_actual > 0
        else f"全 {len(holdings)} 等权 (Positions 无实际买入数据)"
    )

    # 4. join sector
    holdings_df = pd.DataFrame(
        [{"sym": s, "weight": w} for s, w in weights.items()]
    )
    holdings_df = holdings_df.merge(
        ind[["sym", "industry_name"]], on="sym", how="left"
    )
    holdings_df["industry_name"] = holdings_df["industry_name"].fillna("未知")

    sector_w = (
        holdings_df.groupby("industry_name")["weight"]
        .sum()
        .sort_values(ascending=False)
    )
    total = float(sector_w.sum())
    if total <= 0:
        return _placeholder("所有持仓权重为零, 无法计算行业分布")
    sector_pct = sector_w / total * 100.0

    # 5. HHI (Herfindahl-Hirschman Index): sum of squared % shares / 10000 → [0,1]
    hhi = float((sector_pct.values ** 2).sum() / 10000.0)

    # 6. plot — 双图: pie 行业 + bar 个股
    fig, (ax_pie, ax_bar) = plt.subplots(1, 2, figsize=(12, 5))

    colors = plt.cm.Set3(range(len(sector_w)))
    ax_pie.pie(
        sector_pct.values,
        labels=sector_pct.index,
        autopct="%1.1f%%",
        startangle=90,
        colors=colors,
        textprops={"fontsize": 9},
    )
    ax_pie.set_title(f"持仓行业分布 (HHI={hhi:.3f}, n={len(sector_w)} 行业)")

    # bar — 按金额降序的个股, 颜色按所属行业匹配 pie
    sector_to_color = dict(zip(sector_pct.index, colors))
    holdings_sorted = holdings_df.sort_values("weight", ascending=True)
    bar_colors = [sector_to_color.get(ind_name, "#999999")
                  for ind_name in holdings_sorted["industry_name"]]
    from dashboard.utils.stock_names import code_with_name
    bar_labels = [
        f"{code_with_name(row.sym)} ({row.industry_name})"
        for row in holdings_sorted.itertuples()
    ]
    ax_bar.barh(bar_labels, holdings_sorted["weight"], color=bar_colors)
    ax_bar.set_xlabel("Weight (¥)")
    ax_bar.set_title("个股权重 (颜色=所属行业)")
    ax_bar.tick_params(axis="y", labelsize=9)

    fig.tight_layout()
    img_b64 = fig_to_base64(fig)
    plt.close(fig)

    # 7. HHI 警告 + 行业权重表 HTML
    hhi_color, hhi_label = _classify_hhi(hhi)

    table_rows = []
    for ind_name, pct in sector_pct.items():
        amt = float(sector_w[ind_name])
        cell_color = _pct_color(pct)
        table_rows.append(
            f"<tr>"
            f"<td>{html.escape(str(ind_name))}</td>"
            f'<td style="color:{cell_color}; text-align:right; font-weight:500;">{pct:.1f}%</td>'
            f'<td style="text-align:right;">¥{amt:,.0f}</td>'
            f"</tr>"
        )
    table_html = (
        '<table class="data" style="margin-top:12px;">'
        "<thead><tr>"
        "<th>行业 (SW level-1)</th>"
        '<th style="text-align:right;">权重%</th>'
        '<th style="text-align:right;">金额</th>'
        "</tr></thead>"
        f"<tbody>{''.join(table_rows)}</tbody>"
        "</table>"
    )

    summary_html = (
        '<div class="card-explanation" style="margin-top:8px; font-size:13px;">'
        f"HHI = <span style='color:{hhi_color}; font-weight:600;'>{hhi:.3f}</span> "
        f"(<span style='color:{hhi_color};'>{hhi_label}</span>) "
        f"&nbsp;|&nbsp; 阈值: &lt;{HHI_LOW:.2f} 低 / {HHI_LOW:.2f}–{HHI_MID:.2f} 中 / &gt;{HHI_MID:.2f} 高"
        "<br>单行业 &gt;50% 红色, &gt;30% 橙色 警告."
        f"<br>权重基准: {html.escape(weighting_basis)}."
        "</div>"
    )

    return (
        f"<img src='{img_b64}' style='max-width:100%;' alt='Holdings sector breakdown' />"
        f"{summary_html}"
        f"{table_html}"
    )
